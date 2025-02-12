from PySide2.QtUiTools import QUiLoader
from PySide2.QtCore import QCoreApplication, QEvent, QSize, QMetaObject, Qt, SLOT, Slot
from PySide2.QtGui import QBitmap, QColor, QCursor, QIcon, QImage, QKeySequence, QPainter, QPalette, QPixmap, QResizeEvent, QColor
from PySide2.QtWidgets import QApplication, QCheckBox, QComboBox, QDateEdit, QDateTimeEdit, QDial, QDockWidget, QDoubleSpinBox, QFileDialog, QDialog, QFontComboBox, QGraphicsGridLayout, QGraphicsOpacityEffect, QHBoxLayout, QInputDialog, QLCDNumber, QLabel, QLineEdit, QMainWindow, QMenu, QProgressBar, QPushButton, QRadioButton, QScrollArea, QSizePolicy, QSlider, QSpinBox, QStatusBar, QTimeEdit, QToolBar, QGridLayout, QVBoxLayout, QWidget, QAction, QShortcut


from traits.api import HasTraits, Instance, on_trait_change, Range
from traitsui.api import View, Item
from mayavi.core.ui.api import MayaviScene, MlabSceneModel, SceneEditor
from mayavi import mlab

import mayavi
import numpy as np
import os
from PIL import Image, ImageQt
from AnnotationSpace3D import AnnotationSpace3D
import sys
import matplotlib.pyplot as plt
from helpers import read_tiff, create_image_dict, create_colour_array
import asyncio
from openpyxl import Workbook, load_workbook
import math

COLORS = {
    '#ff0000': [255, 0, 0, 255],
    '#35e3e3': [53, 227, 227, 255],
    '#5ebb49': [94, 187, 73, 255],
    '#ffd035': [255, 208, 53, 255],
}

ERASER_COLOR_RGBA = [255, 255, 255, 255]
INIT_COLOR_RGBA = COLORS['#ff0000']

p = 'xy' # xy initially, yz, xz
current_slide = {'xy': 0, 'xz': 0, 'yz': 0}
annot3D = -1
w, h, d = 500, 500, 25

def get_filled_pixmap(pixmap_file):
    pixmap = QPixmap(pixmap_file)
    mask = pixmap.createMaskFromColor(QColor('black'), Qt.MaskOutColor)
    pixmap.fill((QColor('white')))
    pixmap.setMask(mask)
    return pixmap

class Visualization(HasTraits):
    scene = Instance(MlabSceneModel, ())

    @on_trait_change('scene.activated')
    

    def update_plot(self):

        self.image_dictionary=create_image_dict()
        self.current_image_number=0
        self.current_point_index=0
        self.transparancy=1.0

        self.showVolume=True
        self.showResults=False

        self.amount_of_points=20
        self.colour_array=create_colour_array()
        
        self.point_location_data=[[[None]*3 for i in range(self.amount_of_points)]for j in range(len(self.image_dictionary))] #information of point locations in every time step
        
        self.mayavi_result_dots=[None for i in range(len(self.image_dictionary))] #for mayvi to store points for the display view

        self.mayavi_result_lines=[None for i in range(len(self.image_dictionary)-1)] #for mayavi to store lines for the display view

        self.mayavi_dots=[None for i in range(self.amount_of_points)]#location for mayavi to store individual dots

        self.figure = mlab.gcf(engine=self.scene.engine)#nodig voor de picker functie

        global annot3D
        npimages = annot3D.get_npimages()

        self.x_lenght=len(npimages[0][0])
        self.y_lenght=len(npimages[0])
        self.z_lenght=len(npimages)

        self.sphere_size=10

        npspace = annot3D.get_npspace()
        self.npspace_sf = mlab.pipeline.scalar_field(npspace) # scalar field to update later
        self.volume = mlab.pipeline.volume(mlab.pipeline.scalar_field(npimages),color=(0,1,0),vmin=0,vmax=np.amax(npimages)*self.transparancy)

        self.figure.on_mouse_pick(self.picker_callback)

        segmask = mlab.pipeline.iso_surface(self.npspace_sf, color=(1.0, 0.0, 0.0))
        self.scene.background = (0.1, 0.1, 0.1)  
        #mlab.orientation_axes()


    def draw_point(self,new_x,new_y,new_z):#updates point data and draws updated point
        #update point data
        self.point_location_data[self.current_image_number][self.current_point_index]=[new_x,new_y,new_z]
        #draw new point
        if self.mayavi_dots[self.current_point_index] is not None:
            self.mayavi_dots[self.current_point_index].remove()
            self.mayavi_dots[self.current_point_index]=None
        self.mayavi_dots[self.current_point_index]=mlab.points3d(new_x,new_y,new_z,color=self.colour_array[self.current_point_index],scale_factor=self.sphere_size)

    def draw_previous_point(self): #places the point in the same location as it was in the previous image number
        if self.current_image_number==0: #there is no previous point for index 0
            return
        elif self.point_location_data[self.current_image_number-1][self.current_point_index][0]==None: #check if previous dot location is not None
            return
        new_location=self.point_location_data[self.current_image_number-1][self.current_point_index]
        self.draw_point(new_location[0],new_location[1],new_location[2])

    def delete_point(self):
        #update point data
        self.point_location_data[self.current_image_number][self.current_point_index]=[None,None,None]
        #delete point
        if self.mayavi_dots[self.current_point_index] is not None:
            self.mayavi_dots[self.current_point_index].remove()
            self.mayavi_dots[self.current_point_index]=None

    def delete_all_points(self):
        for i in range(self.amount_of_points):
            if self.mayavi_dots[i] is not None:
                self.mayavi_dots[i].remove()
                self.mayavi_dots[i]=None

    def redraw_all_points(self): #redraws all points, used to update points for the next timestep
        self.delete_all_points()
        for i in range(self.amount_of_points):
            if self.point_location_data[self.current_image_number][i][0] is not None: #check if x cordinate is not no to see if a point needs to be placed
                self.mayavi_dots[i]=mlab.points3d(self.point_location_data[self.current_image_number][i][0],self.point_location_data[self.current_image_number][i][1],self.point_location_data[self.current_image_number][i][2],color=self.colour_array[i],scale_factor=self.sphere_size)

    def add_value_to_point(self,added_value):
        old_value=self.point_location_data[self.current_image_number][self.current_point_index]
        if old_value[0] is not None:#check if value exists 
            self.draw_point(old_value[0]+added_value[0],old_value[1]+added_value[1],old_value[2]+added_value[2])

    def update_volume(self,next_or_previous='next'):
        if next_or_previous=="next":
            if self.current_image_number==len(self.image_dictionary)-1:
                return
            self.current_image_number+=1
        elif next_or_previous=="previous":
            if self.current_image_number==0:
                return
            self.current_image_number-=1
        elif isinstance(next_or_previous,int): #used for goto function to go to a specific slide
            if len(self.image_dictionary)-1<next_or_previous:
                self.current_image_number=len(self.image_dictionary)-1
            else:
                self.current_image_number=next_or_previous
        window.load_source_file('data/'+self.image_dictionary[self.current_image_number])
        npimages = annot3D.get_npimages()
        if self.volume is not None:
            self.volume.remove()
            self.npspace_sf.remove()
        npspace = annot3D.get_npspace()
        self.npspace_sf = mlab.pipeline.scalar_field(npspace)
        self.volume = mlab.pipeline.volume(mlab.pipeline.scalar_field(npimages),color=(0,1,0),vmax=np.amax(npimages)*self.transparancy)
        self.redraw_all_points()
        #mlab.orientation_axes()
    
    def remove_volume(self):
        if self.volume is not None: #can't be removed if it isn't there in the first place
            self.volume.remove()
            self.volume=None
            self.npspace_sf.remove()
            self.npspace_sf=None
    
    def draw_results(self):
        for i in range(len(self.point_location_data)):
            if self.point_location_data[i][self.current_point_index][0] is not None: #check if point exists
                x_cordinate=self.point_location_data[i][self.current_point_index][0]
                y_cordinate=self.point_location_data[i][self.current_point_index][1]
                z_cordinate=self.point_location_data[i][self.current_point_index][2]
                self.mayavi_result_dots[i]=mlab.points3d(x_cordinate,y_cordinate,z_cordinate,color=self.colour_array[self.current_point_index],scale_factor=3)


                if i!=0 and self.point_location_data[i-1][self.current_point_index][0]!=None: #check if previous point is not None
                    x_cordinates=[self.point_location_data[i-1][self.current_point_index][0],x_cordinate]
                    y_cordinates=[self.point_location_data[i-1][self.current_point_index][1],y_cordinate]
                    z_cordinates=[self.point_location_data[i-1][self.current_point_index][2],z_cordinate]
                    self.mayavi_result_lines[i-1]=mlab.plot3d(x_cordinates,y_cordinates,z_cordinates,color=(0,0.9,0),tube_radius=1)

    def remove_results(self):
        for i in range(len(self.mayavi_result_dots)):
            if self.mayavi_result_dots[i] is not None:
                self.mayavi_result_dots[i].remove()
                self.mayavi_result_dots[i]=None

        for i in range(len(self.mayavi_result_lines)):
            if self.mayavi_result_lines[i] is not None:
                self.mayavi_result_lines[i].remove()
                self.mayavi_result_lines[i]=None

    def change_result(self): #changes showing and not showing results
        if self.showResults==False:
            self.showResults=True
            self.draw_results()
            window.ToggleVolumeButton.setEnabled(True)
        else:
            self.remove_results()
            self.showResults=False
            window.ToggleVolumeButton.setEnabled(False)
            if self.showVolume==False:
                self.toggle_volume()

    def toggle_volume(self):
        if self.showVolume==True:
            self.showVolume=False
            self.remove_volume()
            self.delete_all_points()
        else:
            self.showVolume=True
            self.redraw_all_points()
            self.update_volume(None)

    def picker_callback(self,picker):
        if self.showVolume==True:
            #print(dir(picker))
            cordinates=picker.pick_position
            self.draw_point(cordinates[0],cordinates[1],cordinates[2])

    def save_data(self,file_name="test_file"):
        point_data_list=[] #meant to put in data from point_location_data that is not None, later used for export to excel
        for i in range(len(self.image_dictionary)):
            for j in range(self.amount_of_points):
                if self.point_location_data[i][j][0] is not None:
                    point_data_list.append([i,j,self.point_location_data[i][j][0],self.point_location_data[i][j][1],self.point_location_data[i][j][2]])

        book=Workbook()
        sheet=book.active
        sheet.append(["timestep","dot","x","y","z"])
        for row in point_data_list:
            sheet.append(row)
        book.save(file_name)

    def export_data(self,file_name,new_x_size,new_y_size,new_z_size):
        x_mod=new_x_size/self.x_lenght #the multiplyer to correct the cordinates to the real size
        y_mod=new_y_size/self.y_lenght
        z_mod=new_z_size/self.z_lenght
        point_data_list=[] #meant to put in data from point_location_data that is not None, later used for export to excel
        for i in range(len(self.image_dictionary)): #for every slice number
            for j in range(self.amount_of_points): #for every tracking number
                if self.point_location_data[i][j][0] is not None:
                    x=self.point_location_data[i][j][0]*x_mod
                    y=self.point_location_data[i][j][1]*y_mod
                    z=self.point_location_data[i][j][2]*z_mod
                    point_data_list.append([j,i,z,y,x,-1]) #the x and z are switched during data import, that is why they are switched back in export
        encounterd_points=[]
        for i in range(len(point_data_list)):
            for j in encounterd_points:
                if point_data_list[i][0]==j[0] and point_data_list[i][1]==j[1]+1:
                    encounterd_points.remove(j)
                    x_exponent=math.pow(point_data_list[i][2]-j[2],2)
                    y_exponent=math.pow(point_data_list[i][3]-j[3],2)
                    z_exponent=math.pow(point_data_list[i][4]-j[4],2)
                    point_data_list[i][5]=math.sqrt(x_exponent+y_exponent+z_exponent)
                    break
            encounterd_points.append(point_data_list[i])

        book=Workbook()
        sheet=book.active
        sheet.append(["tracking number","slice number","x","y","z","Distance"])
        for row in point_data_list:
            sheet.append(row)
        book.save(file_name)

    def load_data(self,file_name="test_file.xlsx"):
        book=load_workbook(filename=file_name)
        sheet=book.active
        self.point_location_data=[[[None]*3 for i in range(self.amount_of_points)]for j in range(len(self.image_dictionary))] #set data back to None to remove old data
        for row in sheet.values:
            if type(row[0])==int: #done to skip the first row that doesn't give data
                self.point_location_data[row[0]][row[1]]=[row[2],row[3],row[4]]
        self.redraw_all_points()
    
    def update_annot(self): # update the scalar field and visualization auto updates
        npspace = annot3D.get_npspace()
        self.npspace_sf.mlab_source.trait_set(scalars=npspace) 

    view = View(Item('scene', editor=SceneEditor(scene_class=MayaviScene), height=250, width=300, show_label=False), resizable=True )

class MayaviQWidget(QWidget):
    def __init__(self, parent=None):
        QWidget.__init__(self, parent)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0,0,0,0)
        layout.setSpacing(0)
        self.visualization = Visualization()

        self.ui = self.visualization.edit_traits(parent=self, kind='subpanel').control
        layout.addWidget(self.ui)
        self.ui.setParent(self)
    
    def update_annot(self):
        self.visualization.update_annot()

class QPaletteButton(QPushButton):
    def __init__(self, color):
        super().__init__()
        self.setFixedSize(QSize(24,24))
        self.color = color
        self.setStyleSheet("background-color: %s;" % color)

class Label(QLabel):

    def __init__(self):
        super(Label, self).__init__()
        self.pixmap_width: int = 1
        self.pixmapHeight: int = 1

    def setPixmap(self, pm: QPixmap) -> None:
        self.pixmap_width = pm.width()
        self.pixmapHeight = pm.height()

        self.updateMargins()
        super(Label, self).setPixmap(pm)

    def resizeEvent(self, a0: QResizeEvent) -> None:
        self.updateMargins()
        super(Label, self).resizeEvent(a0)

    def updateMargins(self):
        if self.pixmap() is None:
            return
        pixmapWidth = self.pixmap().width()
        pixmapHeight = self.pixmap().height()
        if pixmapWidth <= 0 or pixmapHeight <= 0:
            return
        w, h = self.width(), self.height()
        if w <= 0 or h <= 0:
            return

        if w * pixmapHeight > h * pixmapWidth:
            m = int((w - (pixmapWidth * h / pixmapHeight)) / 2)
            self.setContentsMargins(m, 0, m, 0)
        else:
            m = int((h - (pixmapHeight * w / pixmapWidth)) / 2)
            self.setContentsMargins(0, m, 0, m)

class MainWindow(QMainWindow):
    c = {'xy': 0, 'xz': 0, 'yz': 0}
    
    dims = (500, 500, 25) # w, h, d
    slides={}
    plane_depth = {}
    slide_annotations = {}
    num_slides = 0
    npimages = -1

    def load_source_file(self, filename):
        global COLORS, p, current_slide, annot3D

        self.slides['xy'], self.slides['xz'], self.slides['yz'] = read_tiff(filename)

        self.npimages = self.slides['xy']
        
        w = self.npimages[0].shape[0]
        h = self.npimages[0].shape[1]
        d = self.npimages.shape[0]

        annot3D = AnnotationSpace3D(self.npimages, (d, w, h), INIT_COLOR_RGBA)

        self.plane_depth = {
            'xy': d,
            'xz': w,
            'yz': h
        }

        self.dims = (w, h, d)

        self.num_slides = self.plane_depth[p]

        self.slide_annotations = {
            'xy': [] * d, 
            'xz': [] * w, 
            'yz': [] * h
        }

    def __init__(self):
        super().__init__()
        
    # INIT ANNOT LOAD UP
        temp_dict=create_image_dict() #creates a dict so it can load the first file, there might be a beter way to load the first file since this dict is only used once
        self.load_source_file('data/'+temp_dict[0]) #load the first image in the dict

        if len(sys.argv) == 2:
            global annot3D
            print("Set server URL to", sys.argv[1])
            annot3D.set_server_url(sys.argv[1])  

        w = QWidget()
        
        l = QHBoxLayout()
        w.setLayout(l)

    # CANVAS LAYOUT
        canvas_layout = QGridLayout()
        canvas_layout.setAlignment(Qt.AlignLeft)
        sub_canvas_bar_transparancy_layout = QGridLayout()
        sub_canvas_bar_size_layout= QGridLayout()
        sub_canvas_functions_layout=QGridLayout()
        sub_canvas_slide_and_selector_layout=QGridLayout()

    # TOOLBAR, STATUSBAR, MENU
        self.setup_bar_actions()

    # MAYAVI RENDER VIEW
        self.rdock = QDockWidget("Render View", self) # render dock
        self.rdock.setFeatures(self.rdock.features() & ~QDockWidget.DockWidgetClosable) # unclosable
        self.rdock.setAllowedAreas(Qt.LeftDockWidgetArea | Qt.RightDockWidgetArea)
        container = QWidget(self.rdock)
        self.mayavi_widget = MayaviQWidget(container)
        self.rdock.setWidget(self.mayavi_widget)
        self.addDockWidget(Qt.RightDockWidgetArea, self.rdock)
    
    # GENERAL WINDOW PROPS
        self.setWindowTitle("Cell Annotation")
        self.setCentralWidget(w)

        def create_button(self,text,update=[0,0,0]): #function to create standard button
            self.button=QPushButton(text)
            #self.button.setFixedWidth(width)
            self.button.clicked.connect(lambda: self.mayavi_widget.visualization.add_value_to_point(update))
            self.button.setMinimumSize(60,60)
            return self.button    

        self.delete_button=QPushButton('delete')
        self.delete_button.clicked.connect(lambda: self.mayavi_widget.visualization.delete_point())
        self.delete_button.setMinimumSize(50,50)
        sub_canvas_functions_layout.addWidget(self.delete_button,0,1)

        self.create_button=QPushButton('continue')
        self.create_button.clicked.connect(lambda: self.mayavi_widget.visualization.draw_previous_point())
        self.create_button.setMinimumSize(50,50)
        sub_canvas_functions_layout.addWidget(self.create_button,0,0)

        self.result_button=QPushButton("trajectory")
        self.result_button.clicked.connect(lambda: self.mayavi_widget.visualization.change_result())
        self.result_button.setMinimumSize(50,50)
        sub_canvas_functions_layout.addWidget(self.result_button,0,3)

        gotoButton = QPushButton('goto')
        gotoButton.clicked.connect(self.goto_slide)
        gotoButton.setMinimumSize(50,50)
        sub_canvas_slide_and_selector_layout.addWidget(gotoButton,0,0)

        self.x_label = QLabel('X')
        self.x_label.setMinimumSize(50,50)
        self.y_label = QLabel('Y')
        self.y_label.setMinimumSize(50,50)
        self.z_label = QLabel('Z')
        self.z_label.setMinimumSize(50,50)

        self.slide_label = QLabel('slide 1/'+str(len(create_image_dict()))) #number of current slide modified when switching
        self.slide_label.setMinimumSize(50,50)
        sub_canvas_slide_and_selector_layout.addWidget(self.slide_label,0,3)

        canvas_layout.addWidget(self.x_label,0,0)
        canvas_layout.addWidget(self.y_label,1,0)
        canvas_layout.addWidget(self.z_label,2,0)
        

        canvas_layout.addWidget(create_button(self,'-5',[-5,0,0]),0,1) #create x-5 button
        canvas_layout.addWidget(create_button(self,'-1',[-1,0,0]),0,2) #create x-1 button
        canvas_layout.addWidget(create_button(self,'+1',[1,0,0]),0,3) #create x+1 button
        canvas_layout.addWidget(create_button(self,'+5',[5,0,0]),0,4) #create x+5 button

        canvas_layout.addWidget(create_button(self,'-5',[0,-5,0]),1,1) #create y-5 button
        canvas_layout.addWidget(create_button(self,'-1',[0,-1,0]),1,2) #create y-1 button
        canvas_layout.addWidget(create_button(self,'+1',[0,1,0]),1,3) #create y+1 button
        canvas_layout.addWidget(create_button(self,'+5',[0,5,0]),1,4) #create y+5 button

        canvas_layout.addWidget(create_button(self,'-5',[0,0,-5]),2,1) #create z-5 button
        canvas_layout.addWidget(create_button(self,'-1',[0,0,-1]),2,2) #create z-1 button
        canvas_layout.addWidget(create_button(self,'+1',[0,0,1]),2,3) #create z+1 button
        canvas_layout.addWidget(create_button(self,'+5',[0,0,5]),2,4) #create z+5 button

        def change_selected_point(new_point):
            new_point=new_point.split(" ")[1] #split the string and take the number
            self.mayavi_widget.visualization.current_point_index=int(new_point)-1 #-1 becouse index starts at 0
            if self.mayavi_widget.visualization.showResults==True: #change between different results if mode is result
                self.mayavi_widget.visualization.remove_results()
                self.mayavi_widget.visualization.draw_results()

        point_amount=20 #amount of points in pointlist
        point_list=[]   #list for the selectable points in the combobox
        for i in range(point_amount):
            point_list.append("cell "+str(i+1))

        selection_box=QComboBox()
        selection_box.addItems(point_list)    
        selection_box.currentIndexChanged.connect(lambda: change_selected_point(selection_box.currentText()))
        selection_box.setMinimumSize(50,50)
        sub_canvas_functions_layout.addWidget(selection_box,0,2)

        ChangeVolumeNextButton = QPushButton('>')
        ChangeVolumeNextButton.clicked.connect(self.change_volume_model_next)
        ChangeVolumeNextButton.setMinimumSize(50,50)
        sub_canvas_slide_and_selector_layout.addWidget(ChangeVolumeNextButton,0,2)

        ChangeVolumePreviusButton = QPushButton('<')
        ChangeVolumePreviusButton.clicked.connect(self.change_volume_model_previous)
        ChangeVolumePreviusButton.setMinimumSize(50,50)
        sub_canvas_slide_and_selector_layout.addWidget(ChangeVolumePreviusButton,0,1)

        self.ToggleVolumeButton=QPushButton('Volume')
        self.ToggleVolumeButton.clicked.connect(self.mayavi_widget.visualization.toggle_volume)
        self.ToggleVolumeButton.setMinimumSize(50,50)
        sub_canvas_functions_layout.addWidget(self.ToggleVolumeButton,0,4)

        self.ToggleVolumeButton.setEnabled(False)

        self.transparency_slider = QSlider(Qt.Horizontal)
        self.transparency_slider.setValue(10)
        self.transparency_slider.setMinimum(1.0)
        self.transparency_slider.setMaximum(20.0)
        self.transparency_slider.setSingleStep(0.1)
        self.transparency_slider.setFixedWidth(300)
        self.transparency_slider.sliderReleased.connect(self.change_transparancy)

        self.sphere_size_slider = QSlider(Qt.Horizontal)
        self.sphere_size_slider.setValue(10)
        self.sphere_size_slider.setMinimum(1.0)
        self.sphere_size_slider.setMaximum(20.0)
        self.sphere_size_slider.setSingleStep(0.1)
        self.sphere_size_slider.setFixedWidth(300)
        self.sphere_size_slider.sliderReleased.connect(self.change_sphere_size)

        sub_canvas_bar_transparancy_layout.addWidget(QLabel("Transparency"),0,0)
        sub_canvas_bar_transparancy_layout.addWidget(self.transparency_slider,0,1)

        sub_canvas_bar_size_layout.addWidget(QLabel("Sphere size"),0,0)
        sub_canvas_bar_size_layout.addWidget(self.sphere_size_slider,0,1)

        canvas_layout.addLayout(sub_canvas_functions_layout,3,0,1,0,Qt.AlignLeft)
        canvas_layout.addLayout(sub_canvas_slide_and_selector_layout,4,0,1,0,Qt.AlignLeft)
        canvas_layout.addLayout(sub_canvas_bar_transparancy_layout,6,0,1,0,Qt.AlignLeft)
        canvas_layout.addLayout(sub_canvas_bar_size_layout,7,0,1,0,Qt.AlignLeft)

        l.addLayout(canvas_layout)

        #popup layout
        self.popup=QDialog(self)
        self.popup.setWindowTitle("Export information")

        self.popup_layout = QGridLayout()
        x_axis_label = QLabel('X axis size')
        y_axis_label = QLabel('Y axis size')
        z_axis_label = QLabel('Z axis size')

        self.x_input=QSpinBox(maximum=10000,value=self.mayavi_widget.visualization.x_lenght)
        self.y_input=QSpinBox(maximum=10000,value=self.mayavi_widget.visualization.y_lenght)
        self.z_input=QSpinBox(maximum=10000,value=self.mayavi_widget.visualization.z_lenght)

        accept_button=QPushButton('export')
        accept_button.clicked.connect(self.popup.accept)

        cancel_button=QPushButton('cancel')
        cancel_button.clicked.connect(self.popup.reject)


        self.popup_layout.addWidget(x_axis_label,0,0)
        self.popup_layout.addWidget(self.x_input,0,1)
        self.popup_layout.addWidget(y_axis_label,1,0)
        self.popup_layout.addWidget(self.y_input,1,1)
        self.popup_layout.addWidget(z_axis_label,2,0)
        self.popup_layout.addWidget(self.z_input,2,1)
        self.popup_layout.addWidget(accept_button,3,0)
        self.popup_layout.addWidget(cancel_button,3,1)

        self.popup.setLayout(self.popup_layout)

    def setup_bar_actions(self):
        self.statusBar()
        
        exitAction = QAction(QIcon(get_filled_pixmap('graphics/delete.png')), 'Exit', self)
        exitAction.setShortcut(QKeySequence.Quit) # Ctrl+Q
        exitAction.setStatusTip('Exit application')
        exitAction.triggered.connect(self.close)

        loadAnnotAction = QAction(QIcon(get_filled_pixmap('graphics/load.png')), 'Load annotations', self)
        loadAnnotAction.setShortcut(QKeySequence.Open) # Ctrl+O
        loadAnnotAction.setStatusTip('Load new annotations file')
        loadAnnotAction.triggered.connect(self.load_annot_dialog)

        saveAnnotAction = QAction(QIcon(get_filled_pixmap('graphics/save.png')), 'Save annotations', self)
        saveAnnotAction.setShortcut(QKeySequence.Save) # Ctrl+S
        saveAnnotAction.setStatusTip('Save annotations file')
        saveAnnotAction.triggered.connect(self.save_annots_dialog)

        exportAction = QAction(QIcon(get_filled_pixmap('graphics/render.png')), 'Export dataset', self)
        exportAction.setShortcut('Ctrl+E')
        exportAction.setStatusTip('Export source and annotations as dataset directory')
        exportAction.triggered.connect(self.export_dialog)

        gotoAction = QAction('goto', self)
        gotoAction.setShortcut(QKeySequence.Find)
        gotoAction.setStatusTip('Go to specific slide')
        gotoAction.triggered.connect(self.goto_slide)

        ChangeVolumeNextAction = QAction('>', self)
        ChangeVolumeNextAction.setShortcut(QKeySequence.Find)
        ChangeVolumeNextAction.setStatusTip('Go to another volume rendering')
        ChangeVolumeNextAction.triggered.connect(self.change_volume_model_next)

        ChangeVolumePreviusAction = QAction('<', self)
        ChangeVolumePreviusAction.setShortcut(QKeySequence.Find)
        ChangeVolumePreviusAction.setStatusTip('Go to another volume rendering')
        ChangeVolumePreviusAction.triggered.connect(self.change_volume_model_previous)

    # HIDDEN HOTKEY ACTIONS
        slideLeftAction = QAction('Left', self)
        slideLeftAction.setShortcut('Left')
        slideLeftAction.setStatusTip('Slide left')
        slideLeftAction.triggered.connect(self.slide_left)

        slideRightAction = QAction('Right', self)
        slideRightAction.setShortcut('Right')
        slideRightAction.setStatusTip('Slide right')
        slideRightAction.triggered.connect(self.slide_right)

        renderAction = QAction('Render', self)
        renderAction.setShortcut('R')
        renderAction.setStatusTip('Update annotation render')
        renderAction.triggered.connect(self.render)

        self.addAction(slideLeftAction)
        self.addAction(slideRightAction)
        self.addAction(renderAction)
        
    # adding menubar actions 
        menubar = self.menuBar()
        fileMenu = menubar.addMenu('&File')
        fileMenu.addAction(loadAnnotAction)
        fileMenu.addAction(saveAnnotAction)
        fileMenu.addAction(exportAction)
        fileMenu.addAction(exitAction)

    def load_annot_dialog(self):
        fname, _ = QFileDialog.getOpenFileName(self, 'Load annotations file', '.',filter="*.xlsx")

        #global annot3D, current_slide
        if fname:
            window.mayavi_widget.visualization.load_data(fname)
            window.mayavi_widget.visualization.update_volume() #The visualisation needs to be updated after data is loaded
            window.mayavi_widget.visualization.redraw_all_points()
            if window.mayavi_widget.visualization.showResults==True:
                window.mayavi_widget.visualization.change_result()
            

    def save_annots_dialog(self):
        fname, _ = QFileDialog.getSaveFileName(self, 'Save annotations file', '.',"*.xlsx")
        #global annot3D
        if fname:
            window.mayavi_widget.visualization.save_data(fname)  

    def export_dialog(self):
        outcome=self.popup.exec_()
        if outcome==0: #don't do rest of function if cancel button has been pressed
            return
        x_size=self.x_input.value()
        y_size=self.y_input.value()
        z_size=self.z_input.value()
        fname, _ = QFileDialog.getSaveFileName(self, 'Save annotations file', '.',"*.xlsx")
        if fname:
            window.mayavi_widget.visualization.export_data(fname,x_size,y_size,z_size)

    def update_slide_number(self): #used to change slide number display
        slide_number=self.mayavi_widget.visualization.current_image_number
        text="slide "+str(slide_number+1)+"/"+str(len(self.mayavi_widget.visualization.image_dictionary))
        self.slide_label.setText(text)

    def goto_slide(self):
        global p, annot3D

        cs, ok = QInputDialog.getText(self, "Go to slide", "Go to slide")
        if ok and cs.isnumeric(): # current slide cs must be a number
            cs = int(cs)-1
            if cs < 0: # slide out of range
                return
        window.mayavi_widget.visualization.update_volume(cs)
        self.update_slide_number()

    def render(self):
        self.mayavi_widget.update_annot()

    def change_volume_model_next(self):
        if window.mayavi_widget.visualization.showVolume==True:
            window.mayavi_widget.visualization.update_volume('next')
            self.update_slide_number()

    def change_volume_model_previous(self):
        if window.mayavi_widget.visualization.showVolume==True:
            window.mayavi_widget.visualization.update_volume('previous')
            self.update_slide_number()

    def slide_left(self):
        self.change_volume_model_previous()

    def slide_right(self):
        self.change_volume_model_next()

    def change_transparancy(self):
        new_transparancy=self.transparency_slider.value()/10
        window.mayavi_widget.visualization.transparancy=new_transparancy
        window.mayavi_widget.visualization.update_volume(None)

    def change_sphere_size(self):
        new_size=self.sphere_size_slider.value()
        window.mayavi_widget.visualization.sphere_size=new_size
        window.mayavi_widget.visualization.redraw_all_points()

if __name__ == "__main__":
    if not QApplication.instance():
        app = QApplication(sys.argv)
    else:
        app = QApplication.instance()

    app.setStyle('Fusion')
    palette = QPalette()
    palette.setColor(QPalette.Window, QColor(53, 53, 53))
    palette.setColor(QPalette.WindowText, Qt.white)
    palette.setColor(QPalette.Base, QColor(25, 25, 25))
    palette.setColor(QPalette.AlternateBase, QColor(53, 53, 53))
    palette.setColor(QPalette.ToolTipBase, Qt.black)
    palette.setColor(QPalette.ToolTipText, Qt.white)
    palette.setColor(QPalette.Text, Qt.white)
    palette.setColor(QPalette.Button, QColor(53, 53, 53))
    palette.setColor(QPalette.ButtonText, Qt.white)
    palette.setColor(QPalette.BrightText, Qt.red)
    palette.setColor(QPalette.Link, QColor(42, 130, 218))
    palette.setColor(QPalette.Highlight, QColor(42, 130, 218))
    palette.setColor(QPalette.HighlightedText, Qt.black)
    app.setPalette(palette)

    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
